"""导出回测结果到Excel - 详细交易明细和条件检查"""
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from strategy import VolumeBreakoutStrategy
from config import STRATEGY_PARAMS


def export_detailed_trades_to_excel(symbol: str, df: pd.DataFrame, output_file: str = None):
    """
    将单只股票的详细交易信息导出到Excel

    包括：
    - 交易摘要表
    - 详细交易清单
    - 所有信号点（买卖点）的详细条件
    """

    if output_file is None:
        output_file = f'交易详情_{symbol}.xlsx'

    # 计算信号
    strategy = VolumeBreakoutStrategy(STRATEGY_PARAMS)
    signals = strategy.calculate_signals(df)
    trades = strategy.get_trades(df)

    # 创建Excel工作簿
    wb = Workbook()

    # ===== Sheet1: 交易摘要 =====
    ws1 = wb.active
    ws1.title = "交易摘要"

    trades_df = pd.DataFrame(trades) if trades else pd.DataFrame()

    # 添加标题
    ws1['A1'] = f'股票代码: {symbol}'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A3'] = f'数据范围: {df["日期"].min().strftime("%Y-%m-%d")} 到 {df["日期"].max().strftime("%Y-%m-%d")}'

    # 添加统计信息
    row = 5
    ws1[f'A{row}'] = "统计信息"
    ws1[f'A{row}'].font = Font(bold=True, size=12)

    row = 6
    stats = {
        "总交易数": len(trades),
        "总收益率%": trades_df['收益率%'].sum() if len(trades_df) > 0 else 0,
        "平均每笔%": trades_df['收益率%'].mean() if len(trades_df) > 0 else 0,
        "最高收益%": trades_df['收益率%'].max() if len(trades_df) > 0 else 0,
        "最低收益%": trades_df['收益率%'].min() if len(trades_df) > 0 else 0,
    }

    if len(trades_df) > 0:
        wins = len(trades_df[trades_df['收益率%'] > 0])
        stats['胜率%'] = f"{wins/len(trades)*100:.1f}"

    for key, value in stats.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value if isinstance(value, str) else f"{value:.2f}"
        row += 1

    # ===== Sheet2: 详细交易清单 =====
    ws2 = wb.create_sheet("交易清单")

    # 表头
    headers = ['序号', '股票代码', '买入日期', '买入时间', '买入价',
               '卖出日期', '卖出时间', '卖出价', '持有天数', '收益率%', '状态']

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    for idx, trade in enumerate(trades, 1):
        row = idx + 1
        buy_date = trade['买入日期']
        sell_date = trade['卖出日期']

        # 如果是Timestamp类型，转换为字符串
        if hasattr(buy_date, 'strftime'):
            buy_date_str = buy_date.strftime("%Y-%m-%d")
            buy_time_str = "09:30:00"  # 交易时间
        else:
            buy_date_str = str(buy_date)
            buy_time_str = "09:30:00"

        if hasattr(sell_date, 'strftime'):
            sell_date_str = sell_date.strftime("%Y-%m-%d")
            sell_time_str = "14:57:00"  # 收盘时间
        else:
            sell_date_str = str(sell_date)
            sell_time_str = "14:57:00"

        data = [
            idx,
            symbol,
            buy_date_str,
            buy_time_str,
            f"{trade['买入价']:.2f}",
            sell_date_str,
            sell_time_str,
            f"{trade['卖出价']:.2f}",
            trade['持有天数'],
            f"{trade['收益率%']:.2f}",
            trade['状态']
        ]

        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)

            # 收益率着色
            if col == 10:  # 收益率列
                if float(value) > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")

            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 调整列宽
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12
    ws2.column_dimensions['H'].width = 10
    ws2.column_dimensions['I'].width = 10
    ws2.column_dimensions['J'].width = 12
    ws2.column_dimensions['K'].width = 10

    # ===== Sheet3: 所有信号点详情 =====
    ws3 = wb.create_sheet("信号点详情")

    # 获取所有信号点
    signal_points = signals[signals['Buy_Signal'] | signals['Sell_Signal']].copy()

    # 表头
    headers3 = ['日期', '收盘价', 'MA5', 'MA30', '最近3日成交量', '基准20日均量',
                'MA30_Up', 'Volume_Surge', 'MA5_Retest', 'Buy_Signal', 'Sell_Signal']

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充信号点数据
    for idx, (_, row) in enumerate(signal_points.iterrows(), 1):
        ws3.cell(row=idx+1, column=1, value=row['日期'].strftime("%Y-%m-%d"))
        ws3.cell(row=idx+1, column=2, value=f"{row['收盘']:.2f}")
        ws3.cell(row=idx+1, column=3, value=f"{row['MA5']:.2f}" if pd.notna(row['MA5']) else "")
        ws3.cell(row=idx+1, column=4, value=f"{row['MA30']:.2f}" if pd.notna(row['MA30']) else "")
        ws3.cell(row=idx+1, column=5, value=f"{row['Recent3_Vol_Sum']:.0f}" if pd.notna(row.get('Recent3_Vol_Sum')) else "")
        ws3.cell(row=idx+1, column=6, value=f"{row['BaseVol_MA']:.0f}" if pd.notna(row.get('BaseVol_MA')) else "")

        ws3.cell(row=idx+1, column=7, value="是" if row['MA30_Up'] else "否")
        ws3.cell(row=idx+1, column=8, value="是" if row['Volume_Surge'] else "否")
        ws3.cell(row=idx+1, column=9, value="是" if row['MA5_Retest'] else "否")

        # 买卖信号着色
        if row['Buy_Signal']:
            cell = ws3.cell(row=idx+1, column=10, value="BUY")
            cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        if row['Sell_Signal']:
            cell = ws3.cell(row=idx+1, column=11, value="SELL")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

    # 调整列宽
    for col in range(1, 12):
        ws3.column_dimensions[chr(64 + col)].width = 14

    # ===== Sheet4: 策略参数 =====
    ws4 = wb.create_sheet("策略参数")

    ws4['A1'] = "策略参数配置"
    ws4['A1'].font = Font(bold=True, size=12)

    row = 3
    params_info = {
        "MA周期（天）": STRATEGY_PARAMS['ma_period'],
        "量能放大倍数": STRATEGY_PARAMS['volume_multiplier'],
        "5日线周期": STRATEGY_PARAMS['retest_period'],
        "持有天数": STRATEGY_PARAMS['hold_days'],
        "数据范围": f"{df['日期'].min().strftime('%Y-%m-%d')} 到 {df['日期'].max().strftime('%Y-%m-%d')}",
    }

    for key, value in params_info.items():
        ws4[f'A{row}'] = key
        ws4[f'B{row}'] = value
        row += 1

    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 40

    # 保存文件
    wb.save(output_file)
    return output_file


def export_batch_results_to_excel(all_results: dict, index_names: list = None,
                                  output_file: str = "回测结果汇总.xlsx"):
    """
    批量导出多个股票的回测结果到Excel

    all_results: {股票代码: {'trades': [...], 'num_trades': ..., ...}}
    """

    wb = Workbook()

    # ===== Sheet1: 汇总对比 =====
    ws1 = wb.active
    ws1.title = "汇总对比"

    # 表头
    headers = ['股票代码', '交易数', '总收益率%', '平均收益%', '最高收益%', '最低收益%', '胜率%', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # 填充数据
    row = 2
    for symbol, result in all_results.items():
        if result['num_trades'] == 0:
            continue

        trades_df = result['trades_df']

        ws1.cell(row=row, column=1, value=symbol)
        ws1.cell(row=row, column=2, value=result['num_trades'])
        ws1.cell(row=row, column=3, value=f"{result['total_return']:.2f}")
        ws1.cell(row=row, column=4, value=f"{result['avg_return']:.2f}")
        ws1.cell(row=row, column=5, value=f"{trades_df['收益率%'].max():.2f}")
        ws1.cell(row=row, column=6, value=f"{trades_df['收益率%'].min():.2f}")

        wins = len(trades_df[trades_df['收益率%'] > 0])
        win_rate = wins / len(trades_df) * 100 if len(trades_df) > 0 else 0
        ws1.cell(row=row, column=7, value=f"{win_rate:.1f}")
        ws1.cell(row=row, column=8, value="✓ 有交易")

        row += 1

    for col in range(1, 9):
        ws1.column_dimensions[chr(64 + col)].width = 14

    # ===== Sheet2及以后: 每个股票的详细交易 =====
    for symbol, result in all_results.items():
        if result['num_trades'] == 0:
            continue

        ws = wb.create_sheet(f"{symbol}")

        # 表头
        headers = ['序号', '买入日期', '买入价', '卖出日期', '卖出价', '持有天数', '收益率%', '状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        # 填充交易数据
        for idx, trade in enumerate(result['trades'], 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=trade['买入日期'].strftime("%Y-%m-%d"))
            ws.cell(row=idx+1, column=3, value=f"{trade['买入价']:.2f}")
            ws.cell(row=idx+1, column=4, value=trade['卖出日期'].strftime("%Y-%m-%d"))
            ws.cell(row=idx+1, column=5, value=f"{trade['卖出价']:.2f}")
            ws.cell(row=idx+1, column=6, value=trade['持有天数'])

            # 收益率着色
            cell = ws.cell(row=idx+1, column=7, value=f"{trade['收益率%']:.2f}")
            if trade['收益率%'] > 0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws.cell(row=idx+1, column=8, value=trade['状态'])

        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 14

    wb.save(output_file)
    print(f"✓ 已导出到: {output_file}")
    return output_file


if __name__ == "__main__":
    # 测试 - 使用演示数据
    from demo_test_debug import generate_better_mock_data

    symbol = "demo_test"
    df = generate_better_mock_data(symbol)

    output_file = export_detailed_trades_to_excel(symbol, df)
    print(f"✓ 已生成Excel文件: {output_file}")
    print("\nExcel包含以下Sheet:")
    print("  1. 交易摘要 - 统计数据")
    print("  2. 交易清单 - 所有买卖点明细")
    print("  3. 信号点详情 - 所有满足条件的交易点")
    print("  4. 策略参数 - 使用的参数配置")
